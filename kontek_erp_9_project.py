import json
import os
import openpyxl
import logging
import re

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_bom(filepath, project_name):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Check if 'Sheet1' exists
        if 'Sheet1' not in wb.sheetnames:
            raise ValueError("Sheet1 not found")
        sheet = wb['Sheet1']
    except Exception as e:
        logging.error(f"Failed to load {filepath}: {str(e)}")
        return {}, [str(e)]
    
    parts = {}
    errors = []

    # Extract header row to determine column indices dynamically
    headers = {}
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, 1):
        if cell.value:  
            headers[cell.value.strip().upper()] = idx  

    # Define the expected columns, map them to headers if they exist
    try:
        catalog_col = headers['CATALOG']
        manufacturer_col = headers['MANUFACTURE']
        description_col = headers['DESCRIPTION']
        quantity_col = headers.get('QTY')  
        supplier_col = headers.get('SUPPLIER')
    except KeyError as e:
        errors.append(f"Missing expected header: {str(e)}")
        return parts, errors

    for row in range(2, sheet.max_row + 1):
        part_number = sheet.cell(row=row, column=catalog_col).value
        if not part_number:
            continue

        manufacturer = sheet.cell(row=row, column=manufacturer_col).value or 'Unknown'
        description = sheet.cell(row=row, column=description_col).value or 'No Description'

        try:
            raw_quantity = sheet.cell(row=row, column=quantity_col).value if quantity_col else None
            if raw_quantity is not None and re.match(r'^\d+$', str(raw_quantity)):
                quantity = int(raw_quantity)
            else:
                raise ValueError("Invalid quantity")
        except (ValueError, TypeError) as e:
            quantity = 0
            errors.append(f"Invalid quantity for part {part_number} at row {row}: {str(e)}")

        # If part_number exists, update its projects information
        if part_number not in parts:
            parts[part_number] = {
                'description': description,
                'manufacturer': manufacturer,
                'projects': []
            }

        parts[part_number]['projects'].append({
            'project_name': project_name,
            'quantity': quantity
        })

    return parts, errors

def save_json(data, output_filename):
    with open(output_filename, 'w') as f:
        json.dump(data, f, indent=4)

def main():
    directory = "P:/KONTEK/ENGINEERING/ELECTRICAL/Application Development/ERP/9. Electrical BOM Parsing"
    all_parts = {}
    all_errors = {}

    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') or filename.endswith('.xls') and not filename.startswith('~$'):
            filepath = os.path.join(directory, filename)
            logging.info(f"Processing file: {filename}")
            parts, errors = load_bom(filepath, filename)  
            for part_number, part_data in parts.items():
                if part_number not in all_parts:
                    all_parts[part_number] = {
                        'description': part_data['description'],
                        'manufacturer': part_data['manufacturer'],
                        'projects': part_data['projects']
                    }
                else:
                    all_parts[part_number]['projects'].extend(part_data['projects'])
            
            if errors:
                all_errors[filename] = errors
                logging.error(f"Errors encountered in file {filename}: {errors}")

    save_json(all_parts, 'bom.json')
    save_json(all_errors, 'errors.json')

    logging.info("BOM parsing and JSON saving process completed.")

if __name__ == '__main__':
    main()
